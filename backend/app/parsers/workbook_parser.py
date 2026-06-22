"""
Workbook Parser

Reads Excel workbook and converts sheets into domain models.
Structure: Row 1 = Title, Row 2 = Headers, Row 3+ = Data
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import uuid

from app.domain.models import (
    ProjectInfo,
    Resource,
    Sprint,
    WorkItem,
    Dependency,
    Blocker,
    SprintActual,
    ProjectState,
    SkillLevel,
    WorkItemType,
    Priority,
    WorkItemStatus,
    SprintStatus,
    BlockerSeverity,
    BlockerStatus,
    BlockerCategory,
    DependencyType,
)

BLOCKER_CATEGORY_MAP = {
    "vendor": BlockerCategory.VENDOR,
    "hardware": BlockerCategory.HARDWARE,
    "specification": BlockerCategory.SPECIFICATION,
    "resource": BlockerCategory.RESOURCE,
    "environment": BlockerCategory.ENVIRONMENT,
    "security": BlockerCategory.SECURITY,
    "compliance": BlockerCategory.COMPLIANCE,
    "lab issue": BlockerCategory.LAB_ISSUE,
    "hardware / procurement": BlockerCategory.HARDWARE_PROCUREMENT,
    "hardware/procurement": BlockerCategory.HARDWARE_PROCUREMENT,
    "external team dependency": BlockerCategory.EXTERNAL_TEAM_DEPENDENCY,
    "awaiting validation": BlockerCategory.AWAITING_VALIDATION,
    "awaiting validation from central team": BlockerCategory.AWAITING_VALIDATION,
    "tool issue": BlockerCategory.TOOL_ISSUE,
    "license unavailable": BlockerCategory.LICENSE_UNAVAILABLE,
    "license not available": BlockerCategory.LICENSE_UNAVAILABLE,
    "people dependency": BlockerCategory.PEOPLE_DEPENDENCY,
    "approval pending": BlockerCategory.APPROVAL_PENDING,
}


def parse_blocker_category(raw) -> BlockerCategory:
    if not raw:
        return BlockerCategory.OTHER
    return BLOCKER_CATEGORY_MAP.get(str(raw).strip().lower(), BlockerCategory.OTHER)


class WorkbookParseError(Exception):
    """Raised when workbook parsing fails."""
    pass


class WorkbookParser:
    """
    Parses Excel workbook into ProjectState.
    
    Each sheet has:
    - Row 1: Title (skip)
    - Row 2: Column headers
    - Row 3+: Data rows
    """
    
    REQUIRED_SHEETS = [
        "Project_Info", "Team", "Sprint_Plan", "Work_Items",
        "Dependencies", "Blockers", "Sprint_Actuals"
    ]
    
    def __init__(self, file_path: str):
        """Initialize parser with Excel file path."""
        self.file_path = file_path
        self.workbook = None
        self.project_id = str(uuid.uuid4())[:8]
    
    def parse(self) -> ProjectState:
        """
        Parse workbook and return ProjectState.
        
        Returns:
            ProjectState: Parsed and validated project state
            
        Raises:
            WorkbookParseError: If parsing fails
        """
        try:
            # Load cached cell values so formula-derived workbook columns are parsed numerically.
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # Verify all required sheets exist
            self._verify_sheets()
            
            # Parse each sheet
            project_info = self._parse_project_info()
            team = self._parse_team()
            sprints = self._parse_sprints()
            work_items = self._parse_work_items()
            dependencies = self._parse_dependencies()
            blockers = self._parse_blockers()
            actuals = self._parse_sprint_actuals()
            
            # Create and return ProjectState
            return ProjectState(
                project_id=self.project_id,
                project_info=project_info,
                team=team,
                sprints=sprints,
                work_items=work_items,
                dependencies=dependencies,
                blockers=blockers,
                actuals=actuals,
            )
        except Exception as e:
            raise WorkbookParseError(f"Failed to parse workbook: {str(e)}") from e
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _verify_sheets(self) -> None:
        """Verify all required sheets exist."""
        sheet_names = self.workbook.sheetnames
        for sheet in self.REQUIRED_SHEETS:
            if sheet not in sheet_names:
                raise WorkbookParseError(f"Required sheet '{sheet}' not found")
    
    def _get_sheet_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Get data rows from a sheet as list of dicts.
        
        Assumes:
        - Row 1: Title (skip)
        - Row 2: Headers
        - Row 3+: Data
        """
        ws = self.workbook[sheet_name]
        
        # Row 2 contains headers
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(2, col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        if not headers:
            raise WorkbookParseError(f"No headers found in sheet '{sheet_name}' at row 2")
        
        # Row 3+ contains data
        data_rows = []
        for row_idx in range(3, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, start=1):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value is not None:
                    has_data = True
                row_data[header] = cell_value
            
            # Only include rows with at least some data
            if has_data:
                data_rows.append(row_data)
        
        return data_rows
    
    def _parse_project_info(self) -> ProjectInfo:
        """Parse Project_Info sheet (single row)."""
        data_rows = self._get_sheet_data("Project_Info")
        
        if not data_rows:
            raise WorkbookParseError("Project_Info sheet has no data")
        
        row = data_rows[0]
        
        return ProjectInfo(
            project_name=self._get_str(row, "Project Name"),
            sponsor=self._get_str(row, "Sponsor"),
            business_unit=self._get_str(row, "Business Unit"),
            project_manager=self._get_str(row, "Project Manager"),
            start_date=self._get_datetime(row, "Start Date"),
            release_date=self._get_optional_datetime(row, "Release Date"),
            target_end_date=self._get_datetime(row, "Target End Date"),
            sprint_duration_days=self._get_int(row, "Sprint Length (Days)"),
            methodology=self._get_str(row, "Methodology"),
            customer=self._get_str(row, "Customer"),
            status=self._get_str(row, "Status"),
        )
    
    def _parse_team(self) -> List[Resource]:
        """Parse Team sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Team")
        resources = []
        
        for row in data_rows:
            resource_name = self._get_str(row, "Resource Name")
            
            # Generate resource ID from name (use first name initials + last name)
            resource_id = self._generate_resource_id(resource_name)
            
            resources.append(Resource(
                resource_id=resource_id,
                name=resource_name,
                role=self._get_str(row, "Role"),
                primary_skill=self._get_str(row, "Primary Skill"),
                secondary_skill=self._get_optional_str(row, "Secondary Skill"),
                skill_level=self._parse_skill_level(row),
                allocation_pct=self._get_float(row, "Allocation %"),
                availability_pct=self._get_float(row, "Availability %"),
                notes=self._get_optional_str(row, "Notes"),
            ))
        
        return resources
    
    def _parse_sprints(self) -> List[Sprint]:
        """Parse Sprint_Plan sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Sprint_Plan")
        sprints = []
        sprint_number = 0
        
        for row in data_rows:
            # Skip rows that don't have Sprint Name (e.g., summary sections)
            sprint_name = row.get("Sprint Name")
            if not sprint_name or not str(sprint_name).strip().lower().startswith("sprint"):
                continue
            
            sprint_number += 1
            sprint_name = self._get_str(row, "Sprint Name")
            sprint_id = self._generate_sprint_id(sprint_name, sprint_number)
            
            start_date = self._get_datetime(row, "Start Date")
            end_date = self._get_datetime(row, "End Date")
            working_days = self._get_int(row, "Duration (Days)")
            
            sprints.append(Sprint(
                sprint_id=sprint_id,
                sprint_name=sprint_name,
                sprint_number=sprint_number,
                start_date=start_date,
                end_date=end_date,
                working_days=working_days,
                sprint_goal=self._get_str(row, "Sprint Goal"),
                status=self._parse_sprint_status(row),
                planned_velocity_hrs=self._get_float_safe(row, "Velocity (h)"),
                carryover_count=self._get_int(row, "Carry-Over Items"),
            ))
        
        return sprints
    
    def _parse_work_items(self) -> List[WorkItem]:
        """Parse Work_Items sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Work_Items")
        work_items = []
        
        for row in data_rows:
            # Skip rows that don't have Task ID (e.g., summary/totals rows)
            item_id = row.get("Task ID")
            if not item_id or not str(item_id).strip().upper().startswith("WI-"):
                continue
            
            item_id = self._get_str(row, "Task ID")
            
            work_items.append(WorkItem(
                item_id=item_id,
                title=self._get_str(row, "Task Name"),
                work_type=self._parse_work_item_type(row),
                assigned_sprint=self._normalize_sprint_name(self._get_str(row, "Sprint")),
                original_sprint=self._normalize_sprint_name(
                    self._get_optional_str(row, "Orig. Sprint")
                ),
                assigned_resource=self._get_optional_str(row, "Owner"),
                required_skill=self._get_str(row, "Required Skill"),
                priority=self._parse_priority(row),
                estimated_effort_hrs=self._get_float_safe(row, "Orig Est (h)"),
                current_estimate_hrs=self._get_float_safe(row, "Curr Est (h)"),
                actual_effort_hrs=self._get_float_safe(row, "Actual Hrs"),
                remaining_effort_hrs=self._resolve_remaining_effort(
                    row.get("Remaining Hrs"),
                    self._get_float_safe(row, "Curr Est (h)"),
                    self._parse_work_item_status(row),
                ),
                progress_pct=self._get_float_safe(row, "Progress %"),
                status=self._parse_work_item_status(row),
                is_scope_changed=self._parse_yes_no(self._get_optional_str(row, "Scope Change")),
                scope_change_reason=self._get_optional_str(row, "Scope Reason"),
            ))
        
        return work_items
    
    def _resolve_remaining_effort(
        self,
        remaining_value: Any,
        current_estimate: float,
        status: WorkItemStatus,
    ) -> float:
        """Resolve remaining effort based on status and workbook value."""
        if remaining_value is not None:
            if isinstance(remaining_value, str) and remaining_value.strip() == "":
                remaining_value = None
            else:
                try:
                    return float(remaining_value)
                except (ValueError, TypeError):
                    return 0.0

        if status in {WorkItemStatus.DONE, WorkItemStatus.COMPLETED}:
            return 0.0

        if status == WorkItemStatus.NOT_STARTED:
            return current_estimate

        return 0.0

    def _parse_dependencies(self) -> List[Dependency]:
        """Parse Dependencies sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Dependencies")
        dependencies = []
        
        for row in data_rows:
            dependencies.append(Dependency(
                dependency_id=self._get_str(row, "Dep ID"),
                predecessor_item_id=self._get_str(row, "Predecessor Task"),
                successor_item_id=self._get_str(row, "Sucessor Task"),  # Note: typo in sheet
                dependency_type=self._parse_dependency_type(row),
                is_on_critical_path=self._parse_yes_no(self._get_str(row, "Critical Path")),
                lag_days=self._get_int(row, "Lag Days"),
                notes=self._get_optional_str(row, "Notes"),
            ))
        
        return dependencies
    
    def _parse_blockers(self) -> List[Blocker]:
        """Parse Blockers sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Blockers")
        blockers = []
        
        for row in data_rows:
            impacted_str = self._get_str(row, "Impacted Task IDs")
            impacted_ids = [x.strip() for x in impacted_str.split(",")]
            
            # Use notes as description if available, otherwise construct from other fields
            description = self._get_optional_str(row, "Notes")
            if not description:
                description = f"Blocker {self._get_str(row, 'Blocker ID')}: {self._get_str(row, 'Related Task')}"
            
            blockers.append(Blocker(
                blocker_id=self._get_str(row, "Blocker ID"),
                related_item_id=self._get_str(row, "Related Task"),
                impacted_item_ids=impacted_ids,
                description=description,
                severity=self._parse_blocker_severity(row),
                status=self._parse_blocker_status(row),
                owner=self._get_optional_str(row, "Owner"),
                raised_date=self._get_datetime(row, "Raised Date"),
                target_resolution_date=self._get_optional_datetime(row, "Target Resolution"),
                actual_resolution_date=self._get_optional_datetime(row, "Actual Resolution"),
                category=parse_blocker_category(row.get("Category")),
                notes=self._get_optional_str(row, "Notes"),
            ))
        
        return blockers
    
    def _parse_sprint_actuals(self) -> List[SprintActual]:
        """Parse Sprint_Actuals sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Sprint_Actuals")
        actuals = []
        
        for sprint_num, row in enumerate(data_rows, start=1):
            sprint_name = self._get_str(row, "Sprint")
            sprint_id = self._generate_sprint_id(sprint_name, sprint_num)
            
            actuals.append(SprintActual(
                sprint_id=sprint_id,
                sprint_number=sprint_num,
                planned_effort_hrs=self._get_float(row, "Planned Hours"),
                actual_effort_hrs=self._get_float(row, "Actual Hours"),
                variance_hrs=self._get_float_safe(row, "Variance (h)"),
                tasks_planned=self._get_int(row, "Tasks Planned"),
                tasks_completed=self._get_int(row, "Tasks Completed"),
                completion_rate=self._get_float_safe(row, "Completion Rate"),
                carryover_count=self._get_int(row, "Carry-Over Count"),
                scope_change_hours=self._get_float(row, "Scope Change Hours"),
                blocker_impact_hrs=self._get_float(row, "Blocker Impact (h)"),
                notes=self._get_optional_str(row, "Notes"),
            ))
        
        return actuals
    
    # ─── Helper Methods ──────────────────────────────────────────────────────
    
    def _get_str(self, row: Dict, key: str) -> str:
        """Get required string value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        return str(value).strip()
    
    def _get_optional_str(self, row: Dict, key: str) -> Optional[str]:
        """Get optional string value from row."""
        value = row.get(key)
        if value is None:
            return None
        result = str(value).strip()
        return result if result else None
    
    def _get_int(self, row: Dict, key: str) -> int:
        """Get required integer value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        try:
            return int(value)
        except (ValueError, TypeError):
            raise WorkbookParseError(f"Invalid integer for field {key}: {value}")
    
    def _get_float(self, row: Dict, key: str) -> float:
        """Get required float value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        try:
            return float(value)
        except (ValueError, TypeError):
            raise WorkbookParseError(f"Invalid float for field {key}: {value}")
    
    def _get_float_safe(self, row: Dict, key: str) -> float:
        """Get float value, return 0.0 if missing or invalid (for formulas)."""
        value = row.get(key)
        if value is None:
            return 0.0
        if isinstance(value, str) and value.startswith("="):
            # Formula string, can't evaluate
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _get_datetime(self, row: Dict, key: str) -> datetime:
        """Get required datetime value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except (ValueError, TypeError):
            raise WorkbookParseError(f"Invalid datetime for field {key}: {value}")
    
    def _get_optional_datetime(self, row: Dict, key: str) -> Optional[datetime]:
        """Get optional datetime value from row."""
        value = row.get(key)
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except (ValueError, TypeError):
            return None
    
    def _parse_skill_level(self, row: Dict) -> SkillLevel:
        """Parse skill level enum from row."""
        value = self._get_str(row, "Skill Level")
        # Map various spellings to enum
        mapping = {
            "junior": SkillLevel.JUNIOR,
            "intermediate": SkillLevel.INTERMEDIATE,
            "mid": SkillLevel.MID,
            "senior": SkillLevel.SENIOR,
            "advanced": SkillLevel.ADVANCED,
            "expert": SkillLevel.EXPERT,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid skill level: {value}")
        return mapping[key]
    
    def _parse_work_item_type(self, row: Dict) -> WorkItemType:
        """Parse work item type enum from row."""
        value = self._get_str(row, "Type")
        mapping = {
            "feature": WorkItemType.FEATURE,
            "story": WorkItemType.STORY,
            "task": WorkItemType.TASK,
            "bug": WorkItemType.BUG,
            "spike": WorkItemType.SPIKE,
            "defect": WorkItemType.DEFECT,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid work item type: {value}")
        return mapping[key]
    
    def _parse_priority(self, row: Dict) -> Priority:
        """Parse priority enum from row."""
        value = self._get_str(row, "Priority")
        mapping = {
            "critical": Priority.CRITICAL,
            "high": Priority.HIGH,
            "medium": Priority.MEDIUM,
            "low": Priority.LOW,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid priority: {value}")
        return mapping[key]
    
    def _parse_work_item_status(self, row: Dict) -> WorkItemStatus:
        """Parse work item status enum from row."""
        value = self._get_str(row, "Status")
        mapping = {
            "not started": WorkItemStatus.NOT_STARTED,
            "in progress": WorkItemStatus.IN_PROGRESS,
            "done": WorkItemStatus.DONE,
            "completed": WorkItemStatus.COMPLETED,
            "blocked": WorkItemStatus.BLOCKED,
            "spillover": WorkItemStatus.SPILLOVER,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid work item status: {value}")
        return mapping[key]
    
    def _parse_sprint_status(self, row: Dict) -> SprintStatus:
        """Parse sprint status enum from row."""
        value = self._get_str(row, "Status")
        mapping = {
            "not started": SprintStatus.NOT_STARTED,
            "in progress": SprintStatus.IN_PROGRESS,
            "completed": SprintStatus.COMPLETED,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid sprint status: {value}")
        return mapping[key]
    
    def _parse_blocker_severity(self, row: Dict) -> BlockerSeverity:
        """Parse blocker severity enum from row."""
        value = self._get_str(row, "Severity")
        mapping = {
            "critical": BlockerSeverity.CRITICAL,
            "high": BlockerSeverity.HIGH,
            "medium": BlockerSeverity.MEDIUM,
            "low": BlockerSeverity.LOW,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid blocker severity: {value}")
        return mapping[key]
    
    def _parse_blocker_status(self, row: Dict) -> BlockerStatus:
        """Parse blocker status enum from row."""
        value = self._get_str(row, "Status")
        mapping = {
            "open": BlockerStatus.OPEN,
            "resolved": BlockerStatus.RESOLVED,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid blocker status: {value}")
        return mapping[key]
    
    def _parse_dependency_type(self, row: Dict) -> DependencyType:
        """Parse dependency type enum from row."""
        value = self._get_str(row, "Dependency Type")
        mapping = {
            "finish-to-start": DependencyType.FINISH_TO_START,
            "start-to-start": DependencyType.START_TO_START,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid dependency type: {value}")
        return mapping[key]
    
    def _parse_yes_no(self, value: Optional[str]) -> bool:
        """Parse Yes/No string to boolean."""
        if not value:
            return False
        return value.lower() in ["yes", "true", "1", "y"]
    
    def _normalize_sprint_name(self, sprint_name: Optional[str]) -> Optional[str]:
        """Normalize sprint name (e.g., 'Sprint 1' -> 'Sprint 1')."""
        if not sprint_name:
            return None
        return sprint_name.strip()
    
    def _generate_resource_id(self, resource_name: str) -> str:
        """Generate resource ID from name."""
        # Simple approach: use name as-is, could be improved
        return resource_name.lower().replace(" ", "_")
    
    def _generate_sprint_id(self, sprint_name: str, sprint_num: int) -> str:
        """Generate sprint ID from name and number."""
        # Format: SPR-1, SPR-2, etc.
        return f"SPR-{sprint_num}"
