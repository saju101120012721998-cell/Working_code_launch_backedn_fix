from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from app.storage import store
from app.api.models import ApiResponse, ErrorCodes

router = APIRouter(prefix="/api", tags=["Export"])

HIGHLIGHT_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
APPLIED_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")


@router.get("/export")
async def export_workbook(session_id: str = Query(..., description="Session ID")):
    session = store.get_session(session_id)
    if not session:
        raise HTTPException(
            status_code=404,
            detail=ApiResponse(
                success=False,
                error_code=ErrorCodes.SESSION_NOT_FOUND,
                message=f"Session {session_id} not found",
            ).model_dump(),
        )

    workbook_bytes = getattr(session, "workbook_bytes", None)
    if not workbook_bytes:
        raise HTTPException(
            status_code=404,
            detail=ApiResponse(
                success=False,
                error_code=ErrorCodes.SESSION_NOT_FOUND,
                message=f"Workbook bytes not found for session {session_id}",
            ).model_dump(),
        )

    workbook_stream = BytesIO(workbook_bytes)
    wb = load_workbook(workbook_stream)

    # Update Work_Items sheet from session state
    ws = wb["Work_Items"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[2]]
    try:
        task_id_col = headers.index("Task ID") + 1
        curr_est_col = headers.index("Curr Est (h)") + 1
        remaining_col = headers.index("Remaining Hrs") + 1
        scope_reason_col = headers.index("Scope Reason") + 1
        status_col = headers.index("Status") + 1
    except ValueError:
        raise HTTPException(
            status_code=500,
            detail=ApiResponse(
                success=False,
                error_code=ErrorCodes.PROCESSING_ERROR,
                message="Work_Items sheet missing expected columns",
            ).model_dump(),
        )

    item_map = {item.item_id: item for item in session.project_state.work_items}
    for row in ws.iter_rows(min_row=3):
        task_id = row[task_id_col - 1].value
        if not task_id:
            continue
        item = item_map.get(str(task_id).strip())
        if not item:
            continue
        row[curr_est_col - 1].value = item.current_estimate_hrs
        row[curr_est_col - 1].fill = HIGHLIGHT_FILL
        row[remaining_col - 1].value = item.remaining_effort_hrs
        row[remaining_col - 1].fill = HIGHLIGHT_FILL
        row[scope_reason_col - 1].value = item.scope_change_reason
        row[scope_reason_col - 1].fill = HIGHLIGHT_FILL
        row[status_col - 1].value = item.status.value
        row[status_col - 1].fill = HIGHLIGHT_FILL

    # Add or replace Recommended_Actions sheet
    if "Recommended_Actions" in wb.sheetnames:
        wb.remove(wb["Recommended_Actions"])
    actions_ws = wb.create_sheet("Recommended_Actions")

    headers = [
        "Recommendation ID",
        "Type",
        "Action",
        "Target Items",
        "Category",
        "Reason",
        "Escalation Path",
        "Workaround",
        "Delay Reduction (days)",
        "Probability Gain (%)",
        "Effort",
        "Confidence",
        "Status",
        "Date",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = actions_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # No recommendation details available in session directly; leave sheet empty
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    headers = {
        "Content-Disposition": f"attachment; filename= sprint_plan_{session_id}.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(output_stream, headers=headers)
