#!/usr/bin/env python3
"""
Examine workbook structure in detail.
"""

import sys
sys.path.insert(0, '/workspaces/hack_2026_step1/backend')

import openpyxl

workbook_path = '/workspaces/hack_2026_step1/TIO2_Sprint_Intelligence_VALIDATED.xlsx'

wb = openpyxl.load_workbook(workbook_path)

print("=" * 80)
print("WORKBOOK STRUCTURE ANALYSIS")
print("=" * 80)

# List all sheets
print("\nSheet names:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Check each critical sheet
critical_sheets = ["Dependencies", "Blockers", "Sprint_Actuals"]

for sheet_name in critical_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"\n✗ {sheet_name} sheet NOT found!")
        continue
    
    ws = wb[sheet_name]
    print(f"\n[{sheet_name}] Sheet:")
    print(f"  Max rows: {ws.max_row}")
    print(f"  Max columns: {ws.max_column}")
    
    # Check row 1 (title)
    row1_cells = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    print(f"  Row 1 (title): {row1_cells[0] if row1_cells else 'empty'}")
    
    # Check row 2 (headers)
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(2, col).value
        if cell_value:
            headers.append(str(cell_value).strip())
    print(f"  Row 2 (headers): {headers}")
    
    # Count data rows (starting from row 3)
    data_row_count = 0
    for row_idx in range(3, ws.max_row + 1):
        has_data = False
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row_idx, col_idx).value is not None:
                has_data = True
                break
        if has_data:
            data_row_count += 1
    
    print(f"  Data rows (from row 3): {data_row_count}")
    
    # Show first few data rows
    if data_row_count > 0:
        print(f"  First data row (row 3):")
        row_data = {}
        for col_idx, header in enumerate(headers, start=1):
            cell_value = ws.cell(3, col_idx).value
            row_data[header] = cell_value
        for key, val in list(row_data.items())[:3]:
            print(f"    {key}: {val}")

wb.close()

print("\n" + "=" * 80)
